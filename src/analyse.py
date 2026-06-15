"""Analyse de la solution et export du plan de marche (outputs/plan_de_marche.xlsx).

Toutes les fonctions prennent le `Modele` résolu et renvoient des DataFrames
pandas. `exporter_excel` assemble le classeur multi-feuilles.
"""

from __future__ import annotations

import copy
import os

import pandas as pd
import pulp

import equipe
from model import LIGNES, Modele, construire_et_resoudre

TOL = 1e-4


# ---------------------------------------------------------------------------
# Helpers de lecture de la solution
# ---------------------------------------------------------------------------


def _val(v) -> float:
    x = pulp.value(v)
    return 0.0 if x is None else float(x)


def livraison(m: Modele) -> dict[str, float]:
    """cid -> tonnage livré."""
    return {c["id"]: _val(m.x[c["id"]]) for c in m.commandes}


def production(m: Modele) -> dict:
    """(cid, ridx, t) -> tonnage fini produit (>0 seulement)."""
    return {k: _val(v) for k, v in m.p.items() if _val(v) > TOL}


def marge_par_commande(m: Modele) -> dict[str, float]:
    """cid -> marge (CA livré + marges de production)."""
    marges = {c["id"]: c["prix"] * _val(m.x[c["id"]]) for c in m.commandes}
    for (cid, ridx, t), v in m.p.items():
        marges[cid] += m.routes[cid][ridx].marge_unitaire * _val(v)
    return marges


# ---------------------------------------------------------------------------
# Feuilles
# ---------------------------------------------------------------------------


def synthese(m: Modele) -> pd.DataFrame:
    liv = livraison(m)
    demande = sum(c["tonnage"] for c in m.commandes)
    total_livre = sum(liv.values())
    pleines = partielles = refusees = 0
    for c in m.commandes:
        d_, l_ = c["tonnage"], liv[c["id"]]
        if l_ >= d_ - TOL:
            pleines += 1
        elif l_ <= TOL:
            refusees += 1
        else:
            partielles += 1
    lignes = [
        ("Statut", m.statut),
        ("Marge totale (MAD)", round(m.marge, 2)),
        ("Tonnage demandé (T)", round(demande, 1)),
        ("Tonnage livré (T)", round(total_livre, 1)),
        ("Taux de service global (%)", round(100 * total_livre / demande, 2)),
        ("Commandes pleines", pleines),
        ("Commandes partielles", partielles),
        ("Commandes refusées", refusees),
        ("Marge moyenne (MAD/T livrée)", round(m.marge / total_livre, 1) if total_livre else 0),
    ]
    if m.retards_autorises:
        t_retard, t_total = _tonnage_retard(m)
        lignes.append(("Pénalités de retard (MAD)", round(m.penalite_retard, 0)))
        lignes.append(("Tonnage livré en retard (T)", round(t_retard, 1)))
    if m.cout_stockage:
        stock_total = sum(_val(v) for v in m.S.values())
        lignes.append(("Coût de stockage fini (MAD)", round(m.cout_stockage_paye, 0)))
        lignes.append(("Stock fini détenu (T·sem)", round(stock_total, 1)))
    return pd.DataFrame(lignes, columns=["Indicateur", "Valeur"])


def _tonnage_retard(m: Modele) -> tuple[float, float]:
    """(tonnage livré en retard, tonnage total livré)."""
    ech = {c["id"]: c["semaine"] for c in m.commandes}
    retard = total = 0.0
    for (cid, t), v in m.xw.items():
        val = _val(v)
        total += val
        if t > ech[cid]:
            retard += val
    return retard, total


def retards_detail(m: Modele) -> pd.DataFrame:
    """Détail des commandes livrées en retard (vide si modèle de base)."""
    if not m.retards_autorises:
        return pd.DataFrame()
    ech = {c["id"]: c for c in m.commandes}
    rows = []
    for (cid, t), v in m.xw.items():
        val = _val(v)
        c = ech[cid]
        if val > TOL and t > c["semaine"]:
            from model import penalite_priorite
            pen = penalite_priorite(m.donnees, c["priorite"])
            rows.append({
                "ID": cid, "Famille": c["famille"], "Priorité": c["priorite"],
                "Échéance": c["semaine"], "Livré en S": t,
                "Retard (sem)": t - c["semaine"], "Tonnage (T)": round(val, 1),
                "Pénalité (MAD)": round(pen * (t - c["semaine"]) * val, 0),
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Retard (sem)", "Tonnage (T)"], ascending=False)
    return df.reset_index(drop=True)


def _contributions(m: Modele):
    """Itère les contributions (ligne, semaine, famille, θ, flux fini-équiv).

    Compatible Niveau 2 (variables p) et Niveau 3 (variables de segment g)."""
    fam_de = {c["id"]: c["famille"] for c in m.commandes}
    if m.g:  # Niveau 3 : flux par segment
        for (cid, ridx, s, t), v in m.g.items():
            val = _val(v)
            if val <= TOL:
                continue
            rc = m.routes[cid][ridx]
            for ligne in m.segments[(cid, ridx)][s]:
                yield ligne, t, fam_de[cid], rc.theta[ligne], val
    else:    # Niveau 2 : route entière dans la semaine
        for (cid, ridx, t), v in m.p.items():
            val = _val(v)
            if val <= TOL:
                continue
            rc = m.routes[cid][ridx]
            for ligne in rc.processes:
                yield ligne, t, fam_de[cid], rc.theta[ligne], val


def plan_de_marche(m: Modele) -> pd.DataFrame:
    """Tonnage en sortie (output) de chaque ligne × semaine."""
    grille = {ligne: {t: 0.0 for t in m.donnees.semaines} for ligne in LIGNES}
    for ligne, t, fam, theta, val in _contributions(m):
        grille[ligne][t] += theta * val
    df = pd.DataFrame(grille).T
    df.columns = [f"S{t}" for t in m.donnees.semaines]
    df.insert(0, "Ligne", df.index)
    df["Total horizon"] = df[[f"S{t}" for t in m.donnees.semaines]].sum(axis=1)
    return df.round(1).reset_index(drop=True)


def utilisation(m: Modele) -> pd.DataFrame:
    """Taux d'occupation = jours utilisés / jours disponibles."""
    d = m.donnees
    jours_util = {ligne: {t: 0.0 for t in d.semaines} for ligne in LIGNES}
    for ligne, t, fam, theta, val in _contributions(m):
        cad = d.cadences[ligne][fam]
        jours_util[ligne][t] += theta / cad * val

    rows = []
    for ligne in LIGNES:
        util_tot = dispo_tot = 0.0
        row = {"Ligne": ligne}
        for t in d.semaines:
            dispo = d.jours_semaine - d.arrets[ligne][t]
            u = jours_util[ligne][t]
            util_tot += u
            dispo_tot += dispo
            row[f"S{t} (%)"] = round(100 * u / dispo, 1) if dispo else 0.0
        row["Horizon (%)"] = round(100 * util_tot / dispo_tot, 1) if dispo_tot else 0.0
        rows.append(row)
    return pd.DataFrame(rows)


def commandes(m: Modele) -> pd.DataFrame:
    liv = livraison(m)
    rows = []
    for c in m.commandes:
        d_, l_ = c["tonnage"], liv[c["id"]]
        if l_ >= d_ - TOL:
            statut = "Pleine"
        elif l_ <= TOL:
            statut = "Refusée"
        else:
            statut = "Partielle"
        rows.append({
            "ID": c["id"], "Famille": c["famille"], "Grade": c["grade"],
            "Épaisseur": c["epaisseur"], "Largeur": c["largeur"],
            "Semaine": c["semaine"], "Priorité": c["priorite"],
            "Demande (T)": round(d_, 1), "Livré (T)": round(l_, 1),
            "% servi": round(100 * l_ / d_, 1) if d_ else 0.0,
            "Statut": statut,
        })
    return pd.DataFrame(rows)


def _grades_satures(m: Modele) -> set[str]:
    sat = set()
    for g, ctr in m.c_hrc.items():
        if abs(ctr.value()) < 1e-2:  # activité = RHS
            sat.add(g)
    return sat


def _lga_satures(m: Modele) -> set[int]:
    sat = set()
    for (ligne, t), ctr in m.c_cap.items():
        if ligne == "LGA" and abs(ctr.value()) < 1e-2:
            sat.add(t)
    return sat


def refus(m: Modele) -> pd.DataFrame:
    """Commandes non pleinement servies + cause probable de blocage."""
    liv = livraison(m)
    grades_sat = _grades_satures(m)
    lga_sat = _lga_satures(m)
    rows = []
    for c in m.commandes:
        d_, l_ = c["tonnage"], liv[c["id"]]
        if l_ >= d_ - TOL:
            continue
        causes = []
        if c["grade"] in grades_sat:
            causes.append(f"HRC {c['grade']} saturé")
        familles_lga = c["famille"] in ("PPGI", "HDG", "BACR")
        if familles_lga and lga_sat:
            causes.append("LGA saturée")
        if not causes:
            causes.append("Arbitrage marge / capacité")
        rows.append({
            "ID": c["id"], "Famille": c["famille"], "Grade": c["grade"],
            "Demande (T)": round(d_, 1), "Livré (T)": round(l_, 1),
            "% servi": round(100 * l_ / d_, 1) if d_ else 0.0,
            "Contrainte bloquante": " ; ".join(causes),
        })
    return pd.DataFrame(rows)


def shadow_prices(m: Modele) -> pd.DataFrame:
    """Duals des contraintes saturées, triés par valeur décroissante."""
    rows = []
    for (ligne, t), ctr in m.c_cap.items():
        pi = ctr.pi
        if pi is not None and abs(pi) > 1e-6:
            rows.append({"Contrainte": f"Capacité {ligne} S{t}",
                         "Type": "MAD/jour-machine", "Dual": round(pi, 2)})
    for g, ctr in m.c_hrc.items():
        pi = ctr.pi
        if pi is not None and abs(pi) > 1e-6:
            rows.append({"Contrainte": f"HRC {g}",
                         "Type": "MAD/T", "Dual": round(pi, 2)})
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Dual", ascending=False).reset_index(drop=True)
    return df


def marge_famille(m: Modele) -> pd.DataFrame:
    liv = livraison(m)
    mc = marge_par_commande(m)
    rows = []
    for f in m.familles:
        cmds = [c for c in m.commandes if c["famille"] == f]
        demande = sum(c["tonnage"] for c in cmds)
        livre = sum(liv[c["id"]] for c in cmds)
        marge = sum(mc[c["id"]] for c in cmds)
        rows.append({
            "Famille": f,
            "Demande (T)": round(demande, 1),
            "Livré (T)": round(livre, 1),
            "Service (%)": round(100 * livre / demande, 1) if demande else 0.0,
            "Marge (MAD)": round(marge, 0),
            "Marge/T (MAD)": round(marge / livre, 0) if livre > TOL else 0.0,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Vérification empirique des shadow prices HRC
# ---------------------------------------------------------------------------


def verifier_shadow_hrc(m: Modele, delta: float = 5.0) -> pd.DataFrame:
    """Reperturbe la dispo HRC de chaque grade (+delta T) et compare la
    variation de marge observée au dual annoncé par CBC."""
    base = m.marge
    rows = []
    for g in m.c_hrc:
        d2 = copy.deepcopy(m.donnees)
        d2.dispo_hrc[g] += delta
        m2 = construire_et_resoudre(d2)
        dual_obs = (m2.marge - base) / delta
        dual_cbc = m.c_hrc[g].pi or 0.0
        rows.append({
            "Grade": g,
            "Dual CBC (MAD/T)": round(dual_cbc, 1),
            "Dual empirique (MAD/T)": round(dual_obs, 1),
            "Écart": round(abs(dual_cbc - dual_obs), 2),
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------


def _page_de_garde(writer, m: Modele) -> None:
    """Crée une feuille de couverture (logo, titre, équipe, résultats clés)."""
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = writer.book
    ws = wb.create_sheet("Page de garde", 0)
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 28, "C": 28, "D": 28, "E": 28, "F": 3}.items():
        ws.column_dimensions[col].width = w

    navy = "163A6B"
    orange = "E8512E"

    # Logo
    if os.path.exists(equipe.LOGO_PNG):
        img = XLImage(equipe.LOGO_PNG)
        ratio = img.width / img.height
        img.height = 110
        img.width = int(110 * ratio)
        ws.add_image(img, "B2")

    # Titre projet
    ws["B9"] = equipe.TITRE_PROJET
    ws["B9"].font = Font(size=22, bold=True, color=navy)
    ws.merge_cells("B9:E9")
    ws["B10"] = equipe.SOUS_TITRE
    ws["B10"].font = Font(size=12, italic=True, color="3F3F41")
    ws.merge_cells("B10:E10")

    ws["B12"] = equipe.ETABLISSEMENT
    ws["B12"].font = Font(size=11, bold=True, color=orange)
    ws.merge_cells("B12:E12")

    # Équipe
    ws["B14"] = "Équipe projet"
    ws["B14"].font = Font(size=13, bold=True, color=navy)
    for i, membre in enumerate(equipe.EQUIPE):
        c = ws.cell(row=15 + i, column=2, value=f"•  {membre}")
        c.font = Font(size=11)

    # Bandeau résultats clés
    liv = livraison(m)
    demande = sum(c["tonnage"] for c in m.commandes)
    total = sum(liv.values())
    res = [
        ("Statut", m.statut),
        ("Marge totale (MAD)", f"{m.marge:,.0f}"),
        ("Taux de service", f"{100 * total / demande:.1f} %"),
        ("Tonnage livré / demandé (T)", f"{total:,.0f} / {demande:,.0f}"),
    ]
    r0 = 22
    ws.cell(row=r0, column=2, value="Résultats clés").font = Font(
        size=13, bold=True, color=navy)
    for i, (k, v) in enumerate(res):
        rk = ws.cell(row=r0 + 1 + i, column=2, value=k)
        rk.font = Font(size=11, bold=True)
        rk.fill = PatternFill("solid", fgColor="F4F5F7")
        rv = ws.cell(row=r0 + 1 + i, column=3, value=v)
        rv.font = Font(size=11)
        rv.alignment = Alignment(horizontal="left")

    ws.cell(row=r0 + 6, column=2,
            value="Modèle d'optimisation linéaire (PuLP / CBC) — horizon 4 semaines"
            ).font = Font(size=9, italic=True, color="6B7280")


def exporter_excel(m: Modele, chemin: str) -> None:
    """Assemble le classeur plan_de_marche.xlsx (avec page de garde)."""
    feuilles = {
        "Synthèse": synthese(m),
        "Plan de marche": plan_de_marche(m),
        "Utilisation": utilisation(m),
        "Commandes": commandes(m),
        "Refus": refus(m),
        "Shadow_prices": shadow_prices(m),
        "Marge_famille": marge_famille(m),
    }
    with pd.ExcelWriter(chemin, engine="openpyxl") as writer:
        for nom, df in feuilles.items():
            df.to_excel(writer, sheet_name=nom[:31], index=False)
        _page_de_garde(writer, m)
